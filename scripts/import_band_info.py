#!/usr/bin/env python3
"""
Import performer arrival times + band members from an Excel spreadsheet into
the `setLists` Firestore collection.

Spreadsheet shape (per sheet, one sheet per day):
  - Row where column B starts with "Arrival:" → performer header.
    Column A = performer name. Column B = arrival text after "Arrival:".
  - Following rows (both cells non-blank) = members (col A = name, col B = phone).
  - Blank row = end of current performer block.

Upsert behavior (non-destructive):
  - Existing setLists doc with matching performer (trimmed, case-insensitive):
    merge `arrivals[dayKey]` + union `members` by name. Leaves `songs`,
    `stage`, etc. untouched.
  - No match: create new doc with only performer / stage="" / arrivals / members
    / songs=[] / performanceOverrides={}.

Usage:
    cd scripts
    python3 import_band_info.py /path/to/spreadsheet.xlsx
    # (prompts for confirmation before writing)
    python3 import_band_info.py /path/to/spreadsheet.xlsx --yes   # non-interactive
"""

import argparse
import sys
from datetime import datetime

import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl


SHEET_DAY_MAP = {
    'thursday': 'thursday',
    'friday':   'friday',
    'saturday': 'saturday',
    'sunday':   'sunday',
}

ALL_DAYS = ['thursday', 'friday', 'saturday', 'sunday']


def initialize_firebase(cred_path='../config/firebase-credentials.json'):
    if not firebase_admin._apps:
        cred = credentials.Certificate(cred_path)
        firebase_admin.initialize_app(cred)
    return firestore.client()


def sheet_to_day_key(sheet_name):
    """Map a sheet name like 'Friday ' to the day key 'friday'. Returns None if no match."""
    key = (sheet_name or '').strip().lower()
    return SHEET_DAY_MAP.get(key)


def parse_sheet(ws):
    """Yield {performer, arrival, members: [{name, phone}]} dicts from one sheet."""
    current = None
    for row in ws.iter_rows(values_only=True):
        col_a = (row[0] if len(row) > 0 else None)
        col_b = (row[1] if len(row) > 1 else None)
        a = str(col_a).strip() if col_a is not None else ''
        b = str(col_b).strip() if col_b is not None else ''

        if not a and not b:
            # Blank row — end of current block
            if current:
                yield current
                current = None
            continue

        if b.lower().startswith('arrival:'):
            # Emit the previous block before starting a new one
            if current:
                yield current
            arrival_text = b[len('arrival:'):].strip()
            current = {'performer': a, 'arrival': arrival_text, 'members': []}
        else:
            if current is None:
                # Orphan row without a header — skip
                continue
            # Member row: name in A, phone in B (phone may be blank)
            if a:
                current['members'].append({'name': a, 'phone': b})

    if current:
        yield current


def find_existing_by_performer(db, performer_name):
    """Return existing setLists doc (id, data) whose performer matches case-insensitively, or None."""
    norm = (performer_name or '').strip().lower()
    for doc in db.collection('setLists').stream():
        d = doc.to_dict() or {}
        existing = str(d.get('performer') or '').strip().lower()
        if existing == norm:
            return doc.id, d
    return None


def merge_members(existing_members, new_members):
    """Union by name (case-insensitive, trimmed). Preserve phone from existing when both present."""
    out = list(existing_members or [])
    seen = {(str(m.get('name') or '').strip().lower()) for m in out if m.get('name')}
    added = 0
    for m in new_members or []:
        name = str(m.get('name') or '').strip()
        if not name:
            continue
        key = name.lower()
        if key in seen:
            continue
        out.append({'name': name, 'phone': str(m.get('phone') or '').strip()})
        seen.add(key)
        added += 1
    return out, added


def plan_upserts(db, spreadsheet_path):
    wb = openpyxl.load_workbook(spreadsheet_path, data_only=True)
    # Aggregate per-performer across all sheets
    aggregated = {}  # key = lowercase performer, value = {performer, arrivals: {}, members: []}
    for sheet_name in wb.sheetnames:
        day_key = sheet_to_day_key(sheet_name)
        if not day_key:
            print(f"  ! Skipping sheet '{sheet_name}' (not a recognized day)")
            continue
        ws = wb[sheet_name]
        for block in parse_sheet(ws):
            perf = block['performer'].strip()
            key = perf.lower()
            if key not in aggregated:
                aggregated[key] = {
                    'performer': perf,
                    'arrivals': {},
                    'members_per_day': [],
                }
            aggregated[key]['arrivals'][day_key] = block['arrival']
            aggregated[key]['members_per_day'].extend(block['members'])

    # Now build a plan: per performer, show what will change
    plan = []
    for key, agg in aggregated.items():
        match = find_existing_by_performer(db, agg['performer'])
        # Deduplicate members collected from multiple days
        seen = set()
        deduped = []
        for m in agg['members_per_day']:
            n = m['name'].strip()
            if not n:
                continue
            nk = n.lower()
            if nk in seen:
                continue
            seen.add(nk)
            deduped.append({'name': n, 'phone': m['phone'].strip()})

        if match:
            doc_id, existing = match
            merged_members, added = merge_members(existing.get('members') or [], deduped)
            new_arrivals = dict(existing.get('arrivals') or {})
            arrival_changes = []
            for day, arrival in agg['arrivals'].items():
                if (new_arrivals.get(day) or '') != arrival:
                    new_arrivals[day] = arrival
                    arrival_changes.append(day)
            plan.append({
                'action': 'UPDATE',
                'doc_id': doc_id,
                'performer': agg['performer'],
                'existing_performer': existing.get('performer'),
                'new_arrivals': new_arrivals,
                'arrival_changes': arrival_changes,
                'merged_members': merged_members,
                'members_added': added,
            })
        else:
            plan.append({
                'action': 'CREATE',
                'doc_id': None,
                'performer': agg['performer'],
                'new_arrivals': agg['arrivals'],
                'merged_members': deduped,
            })

    return plan


def print_plan(plan):
    print(f"\nPlanned changes: {len(plan)} performer(s)")
    print('-' * 60)
    for p in plan:
        if p['action'] == 'CREATE':
            days = ', '.join(f"{d}={p['new_arrivals'][d]}" for d in p['new_arrivals'])
            print(f"  CREATE  {p['performer']}")
            print(f"          arrivals: {days}")
            print(f"          members: {len(p['merged_members'])}")
        else:
            days = ', '.join(p['arrival_changes']) if p['arrival_changes'] else '(no changes)'
            print(f"  UPDATE  {p['performer']}  (doc: {p['doc_id']})")
            print(f"          arrivals touched: {days}")
            print(f"          +{p['members_added']} member(s); total now {len(p['merged_members'])}")
    print('-' * 60)


def apply_plan(db, plan):
    # Batch writes in chunks of 450 (Firestore batch limit is 500)
    batch = db.batch()
    write_count = 0

    for p in plan:
        doc_ref = (db.collection('setLists').document(p['doc_id'])
                   if p['doc_id']
                   else db.collection('setLists').document())
        if p['action'] == 'CREATE':
            data = {
                'performer': p['performer'],
                'stage': '',
                'arrivals': p['new_arrivals'],
                'performanceOverrides': {d: '' for d in ALL_DAYS},
                'members': p['merged_members'],
                'songs': [],
                'createdAt': firestore.SERVER_TIMESTAMP,
                'updatedAt': firestore.SERVER_TIMESTAMP,
            }
            batch.set(doc_ref, data)
        else:
            data = {
                'arrivals': p['new_arrivals'],
                'members': p['merged_members'],
                'updatedAt': firestore.SERVER_TIMESTAMP,
            }
            batch.update(doc_ref, data)

        write_count += 1
        if write_count % 450 == 0:
            batch.commit()
            batch = db.batch()

    if write_count % 450 != 0:
        batch.commit()

    print(f"\n✓ Committed {write_count} write(s).")


def main():
    parser = argparse.ArgumentParser(description='Import performer arrivals + members from Excel.')
    parser.add_argument('spreadsheet', help='Path to the .xlsx file')
    parser.add_argument('--yes', '-y', action='store_true', help='Skip confirmation')
    parser.add_argument('--cred', default='../config/firebase-credentials.json',
                        help='Path to Firebase credentials JSON')
    args = parser.parse_args()

    db = initialize_firebase(args.cred)
    plan = plan_upserts(db, args.spreadsheet)

    if not plan:
        print('No performer data found in spreadsheet.')
        sys.exit(0)

    print_plan(plan)

    if not args.yes:
        reply = input('\nProceed with these writes? [y/N]: ').strip().lower()
        if reply not in ('y', 'yes'):
            print('Aborted — no changes made.')
            sys.exit(0)

    apply_plan(db, plan)


if __name__ == '__main__':
    main()
