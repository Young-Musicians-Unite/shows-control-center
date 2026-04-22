#!/usr/bin/env python3
"""
One-time import of performer arrival times + band members from
../source-data/Untitled spreadsheet (1).xlsx with explicit name mappings
decided by the user:

  - "4tune" on sheet → merge into existing "Fourtune" doc
  - "Rock Ensemble" on sheet → merge into BOTH "Miami Beach Rock Ensemble (Set 1)"
    and "(Set 2)" docs
  - "Homestead Senior High" + "West Little River" on sheet → put BOTH arrivals
    on the existing combined school doc, labeled ("Homestead 1:30pm; WLR 2:00pm")
  - Everything else: case-insensitive trimmed match to existing docs; create
    new docs if no match

Run once with no args to print the plan. Run with --apply to write.
"""

import argparse
import sys
from collections import defaultdict

import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl


SPREADSHEET = '../source-data/Untitled spreadsheet (1).xlsx'
CRED = '../config/firebase-credentials.json'

SHEET_DAY_MAP = {'thursday': 'thursday', 'friday': 'friday',
                 'saturday': 'saturday', 'sunday': 'sunday'}
ALL_DAYS = ['thursday', 'friday', 'saturday', 'sunday']

# Explicit name mappings (keys are normalized = lowercased + trimmed)
MAPPINGS = {
    '4tune': {'type': 'redirect', 'to_performer': 'Fourtune'},
    'rock ensemble': {'type': 'multi_redirect',
                      'to_performers': ['Miami Beach Rock Ensemble (Set 1)',
                                        'Miami Beach Rock Ensemble (Set 2)']},
    'homestead senior high': {'type': 'labeled',
                              'to_performer': 'West Little River K-8 + Homestead Senior High Flagettes Marching Band',
                              'label': 'Homestead'},
    'west little river': {'type': 'labeled',
                          'to_performer': 'West Little River K-8 + Homestead Senior High Flagettes Marching Band',
                          'label': 'WLR'},
}


def initialize_firebase():
    if not firebase_admin._apps:
        cred = credentials.Certificate(CRED)
        firebase_admin.initialize_app(cred)
    return firestore.client()


def normalize(s):
    return (s or '').strip().lower()


def parse_sheet(ws):
    current = None
    for row in ws.iter_rows(values_only=True):
        a = (str(row[0]).strip() if len(row) > 0 and row[0] is not None else '')
        b = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else '')
        if not a and not b:
            if current:
                yield current
                current = None
            continue
        if b.lower().startswith('arrival:'):
            if current:
                yield current
            current = {'performer': a, 'arrival': b[len('arrival:'):].strip(),
                       'members': []}
        else:
            if current is None:
                continue
            if a:
                current['members'].append({'name': a, 'phone': b})
    if current:
        yield current


def build_plan(db, spreadsheet_path):
    # Load all existing setLists into memory
    existing_docs = {}        # doc_id → data
    name_to_doc_id = {}       # normalized performer name → doc_id
    for doc in db.collection('setLists').stream():
        d = doc.to_dict() or {}
        existing_docs[doc.id] = d
        key = normalize(d.get('performer'))
        if key:
            # If duplicates, last one wins for lookup — shouldn't matter here
            name_to_doc_id[key] = doc.id

    # Target-doc → aggregated changes
    # Each target gets: arrivals {day: str}, members [{name, phone}],
    # labeled_arrivals {day: [(label, arrival), ...]}
    targets = defaultdict(lambda: {
        'arrivals': {},
        'members_by_name': {},  # normalized name → {name, phone}
        'labeled_arrivals': defaultdict(list),
    })

    # Performers with no existing match — will be CREATEd as new docs
    # Keyed by the sheet's performer name (trimmed)
    new_docs = defaultdict(lambda: {
        'performer': None,
        'arrivals': {},
        'members_by_name': {},
    })

    wb = openpyxl.load_workbook(spreadsheet_path, data_only=True)
    unmatched = []

    for sheet_name in wb.sheetnames:
        day_key = SHEET_DAY_MAP.get(normalize(sheet_name))
        if not day_key:
            print(f'  skipping sheet: {sheet_name!r}')
            continue
        ws = wb[sheet_name]
        for block in parse_sheet(ws):
            perf = block['performer'].strip()
            key = normalize(perf)
            arrival = block['arrival']
            members = block['members']

            if key in MAPPINGS:
                m = MAPPINGS[key]
                if m['type'] == 'redirect':
                    tgt_id = name_to_doc_id.get(normalize(m['to_performer']))
                    if not tgt_id:
                        unmatched.append(f"redirect target not found: {m['to_performer']}")
                        continue
                    _apply_to_target(targets[tgt_id], day_key, arrival, members)
                elif m['type'] == 'multi_redirect':
                    for tp in m['to_performers']:
                        tgt_id = name_to_doc_id.get(normalize(tp))
                        if not tgt_id:
                            unmatched.append(f"multi_redirect target not found: {tp}")
                            continue
                        _apply_to_target(targets[tgt_id], day_key, arrival, members)
                elif m['type'] == 'labeled':
                    tgt_id = name_to_doc_id.get(normalize(m['to_performer']))
                    if not tgt_id:
                        unmatched.append(f"labeled target not found: {m['to_performer']}")
                        continue
                    targets[tgt_id]['labeled_arrivals'][day_key].append(
                        (m['label'], arrival))
                    # Still merge any members (schools in this sheet have none)
                    for m_ in members:
                        nk = normalize(m_['name'])
                        if nk and nk not in targets[tgt_id]['members_by_name']:
                            targets[tgt_id]['members_by_name'][nk] = {
                                'name': m_['name'].strip(),
                                'phone': (m_.get('phone') or '').strip(),
                            }
            else:
                # Default: match by normalized performer name
                tgt_id = name_to_doc_id.get(key)
                if tgt_id:
                    _apply_to_target(targets[tgt_id], day_key, arrival, members)
                else:
                    # New doc — aggregate across days
                    entry = new_docs[key]
                    entry['performer'] = perf
                    entry['arrivals'][day_key] = arrival
                    for m_ in members:
                        nk = normalize(m_['name'])
                        if nk and nk not in entry['members_by_name']:
                            entry['members_by_name'][nk] = {
                                'name': m_['name'].strip(),
                                'phone': (m_.get('phone') or '').strip(),
                            }

    # Resolve labeled arrivals into final arrival strings (existing doc's
    # arrivals[day] gets OVERWRITTEN with the combined labeled text).
    for tgt_id, t in targets.items():
        for day, entries in t['labeled_arrivals'].items():
            t['arrivals'][day] = '; '.join(f'{lab} {arr}' for lab, arr in entries)

    return existing_docs, targets, new_docs, unmatched


def _apply_to_target(tgt, day_key, arrival, members):
    tgt['arrivals'][day_key] = arrival
    for m in members:
        nk = normalize(m['name'])
        if nk and nk not in tgt['members_by_name']:
            tgt['members_by_name'][nk] = {
                'name': m['name'].strip(),
                'phone': (m.get('phone') or '').strip(),
            }


def union_members(existing_list, new_by_name):
    """Return (final_list, added_count) given existing list and new_by_name dict."""
    out = list(existing_list or [])
    seen = {normalize(m.get('name')) for m in out if m.get('name')}
    added = 0
    for nk, m in new_by_name.items():
        if nk in seen:
            continue
        out.append(m)
        seen.add(nk)
        added += 1
    return out, added


def print_plan(existing_docs, targets, new_docs, unmatched):
    print(f'\n=== UPDATE existing docs ({len(targets)}) ===')
    for tgt_id, t in sorted(targets.items(),
                            key=lambda kv: (existing_docs[kv[0]].get('performer') or '').lower()):
        doc = existing_docs[tgt_id]
        name = doc.get('performer') or '(unnamed)'
        print(f'\n  "{name}" [{tgt_id}]')
        for day in ALL_DAYS:
            if day in t['arrivals']:
                before = (doc.get('arrivals') or {}).get(day) or '(blank)'
                print(f'    arrivals.{day}: {before!r} → {t["arrivals"][day]!r}')
        merged, added = union_members(doc.get('members') or [], t['members_by_name'])
        print(f'    members: {len(doc.get("members") or [])} existing + {added} new = {len(merged)}')

    print(f'\n=== CREATE new docs ({len(new_docs)}) ===')
    for key, entry in sorted(new_docs.items()):
        print(f'\n  "{entry["performer"]}"')
        for day in ALL_DAYS:
            if day in entry['arrivals']:
                print(f'    arrivals.{day}: {entry["arrivals"][day]!r}')
        print(f'    members: {len(entry["members_by_name"])}')

    if unmatched:
        print(f'\n=== WARNINGS ===')
        for u in unmatched:
            print(f'  ! {u}')


def apply_writes(db, existing_docs, targets, new_docs):
    batch = db.batch()
    count = 0

    for tgt_id, t in targets.items():
        doc = existing_docs[tgt_id]
        # Merge arrivals: existing + new (new overrides per day key)
        final_arrivals = dict(doc.get('arrivals') or {})
        final_arrivals.update(t['arrivals'])
        final_members, _ = union_members(doc.get('members') or [], t['members_by_name'])
        batch.update(db.collection('setLists').document(tgt_id), {
            'arrivals': final_arrivals,
            'members': final_members,
            'updatedAt': firestore.SERVER_TIMESTAMP,
        })
        count += 1

    for key, entry in new_docs.items():
        new_ref = db.collection('setLists').document()
        batch.set(new_ref, {
            'performer': entry['performer'],
            'stage': '',
            'arrivals': entry['arrivals'],
            'performanceOverrides': {d: '' for d in ALL_DAYS},
            'members': list(entry['members_by_name'].values()),
            'songs': [],
            'createdAt': firestore.SERVER_TIMESTAMP,
            'updatedAt': firestore.SERVER_TIMESTAMP,
        })
        count += 1

    batch.commit()
    print(f'\n✓ Wrote {count} doc(s).')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--apply', action='store_true',
                        help='Actually write to Firestore (otherwise dry-run)')
    parser.add_argument('--path', default=SPREADSHEET)
    args = parser.parse_args()

    db = initialize_firebase()
    existing_docs, targets, new_docs, unmatched = build_plan(db, args.path)

    print_plan(existing_docs, targets, new_docs, unmatched)

    if args.apply:
        apply_writes(db, existing_docs, targets, new_docs)
    else:
        print('\n(Dry run. Re-run with --apply to write.)')


if __name__ == '__main__':
    main()
